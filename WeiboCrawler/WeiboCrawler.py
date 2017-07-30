# -*- coding: utf-8 -*-
"""
Created on  2017/6/20
Author:     Haiyu Zhu
E-mail:     zhuhaiyu1991@163.com
"""
import os
import urllib.parse
import urllib.request
import urllib.error
import http.cookiejar
import time
import datetime
from lxml import etree
import random
import yaml
import re
import openpyxl
import WeiboLogin
import GlobalVariable as G
from selenium import webdriver
from SearchVerificationCode import verify


class Crawler():
    """
    Crawler class
    """

    def __init__(self):
        self.__url = ""
        # type: people, media, original, etc.
        self.__weibo_type = ""
        # type of weibo, text, pic., video, etc.
        self.__weibo_containing = ""
        self.__cookiejar = None
        # filter data
        self.filter = {}
        self.__keyword = ""
        self.__keyword_encode = ""
        self.__start_time = None
        self.__start_time_curr = None
        self.__stop_time = None
        # time scope of every search, default is 1 hour
        self.__intelval_hour = 1
        self.__time_scope = ""
        self.__region = ""
        self.__subregion = ""
        self.__work_dir = os.getcwd()
        self.__file_name = ""
        self.__is_robot = False
        self.__max_try_num = 4
        # proxy pool
        self.__proxy_pool = []
        # yun da ma
        self.__ydm_username = ""
        self.__ydm_password = ""
        self.__ydm_app_id = 0
        self.__ydm_app_key = ""

    def set_search_target(self, keyword, weibo_type, weibo_containing,
                          start_time, stop_time, region, subregion):
        """
        set necessary information for Weibo advanced search
        """
        self.__keyword = keyword
        self.__encode_keyword()
        self.__weibo_type = weibo_type
        self.__weibo_containing = weibo_containing
        self.__start_time = datetime.datetime.strptime(
            start_time, "%Y-%m-%d-%H")
        self.__start_time_curr = self.__start_time
        self.__stop_time = datetime.datetime.strptime(
            stop_time, "%Y-%m-%d-%H")
        self.__region = region
        self.__subregion = subregion

    def __get_url(self):
        """(WeiboCrawler) -> str
        set the begin url, default is: begin_url=http://s.weibo.com/weibo/
        """
        base_url = "http://s.weibo.com/weibo/"
        region_prefix = "&region=custom:"
        if self.__region == '省/直辖市':
            region_prefix = ""
        self.__url = base_url + self.__keyword_encode + \
            region_prefix + G.districts[self.__region][0] + ":" + \
            G.districts[self.__region][1][self.__subregion] + \
            G.weibo_type[self.__weibo_type] + \
            G.weibo_containing[self.__weibo_containing] + \
            "&timescope=custom:" + self.__time_scope + \
            "&page="
        return self.__url

    def set_cookiejar(self, cookiejar_obj):
        """(WeiboCrawler, http.cookiejar)
        set the cookie because Sina needs users to login
        """
        if None is not cookiejar_obj:
            self.__cookiejar = cookiejar_obj
            cookie_support = urllib.request.HTTPCookieProcessor(cookiejar_obj)
            opener = urllib.request.build_opener(cookie_support)
            urllib.request.install_opener(opener)

    def __encode_keyword(self):
        """
        keyword is urlencoded twice
        """
        keyword = self.__keyword.encode("utf-8")
        once = urllib.parse.urlencode({"kw": keyword})[3:]
        self.__keyword_encode = urllib.parse.urlencode({"kw": once})[3:]

    def __set_search_time_scope(self, start_time, stop_time):
        """(WeiboCrawler, datetime, datetime) -> None
        set the time scope for search, yyyy-mm-dd-hh
        """
        if start_time is None or stop_time is None:
            self.__time_scope == ""
        else:
            self.__time_scope = start_time.strftime(
                "%Y-%m-%d-%H") + ":" + stop_time.strftime("%Y-%m-%d-%H")

    def next_time_scope(self):
        """(WeiboCrawler) -> boolean
        interval is 1 hour
        :return False means stop downloading:
        """
        if self.__start_time_curr >= self.__stop_time:
            return False
        else:
            time_delt = datetime.timedelta(hours=self.__intelval_hour)
            time_next = self.__start_time_curr + time_delt
            self.__set_search_time_scope(self.__start_time_curr, time_next)
            self.__start_time_curr = time_next
            return True

    def set_save_file_name(self, file_name):
        """(WeiboCrawler, str) -> None
        set the name of the file to save data
        """
        if '.xlsx' not in file_name:
            file_name += '.xlsx'
        self.__file_name = file_name

    def set_yun_da_ma(self, app_id, app_key, username, password):
        """
        """
        self.__ydm_app_id = app_id
        self.__ydm_app_key = app_key
        self.__ydm_username = username
        self.__ydm_password = password

    def craw_data(self, buffer):
        """(WeiboCrawler, [(str, str)]) -> boolean
        buffer: [(uname, text_time, text)]
        """
        page_num = 1
        while page_num < 51:
            respondence = ""
            source_url = self.__get_url() + str(page_num)
            if page_num == 1:
                referer = ""
            else:
                referer = self.__get_url() + str(page_num - 1)
            headers = {
                "Host":
                    "s.weibo.com",
                "Referer":
                    referer,
                "User-Agent":
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36"
            }
            req = urllib.request.Request(url=source_url, headers=headers)
            # print(source_url)
            try:
                html = urllib.request.urlopen(req)
                respondence = html.read()
            except urllib.error.URLError as e:
                # show_message(e)
                if e.code == 500:
                    print("http error code: %d" % e.code)
                    return True
                else:
                    print("http error code: %d" % e.code)
                    return False

            lines = respondence.splitlines()
            is_caught = True
            has_more = True
            for line in lines:
                if line.startswith(b'<script>STK && STK.pageletM && STK.pageletM.view({"pid":"pl_weibo_direct"'):
                    is_caught = False
                    idx = line.find(b'html":"')
                    if idx > 0:
                        emoji = re.compile(u'[\uD800-\uDBFF][\uDC00-\uDFFF]')
                        # delete emoji
                        j = str(
                            line[idx + 7:-12], encoding='unicode_escape').replace("\\", "")
                        j = emoji.sub(u"", j)
                        if j.find(u'<div class="search_noresult">') > 0:
                            # no more pages
                            has_more = False
                        else:
                            self.__parse_keyword_search_result(j, buffer)
                        break
            if is_caught:
                self.show_message("Be Caught! Please input verification code in your browser!",
                                  "r")
                self.show_message(
                    "If you have input the verification code, please go on.", "r")
                verify(self.__ydm_app_id, self.__ydm_app_key,
                       self.__ydm_username, self.__ydm_password)
                time.sleep(3)
                return False
                # break
            sleep_time = random.randint(0, 4)
            self.show_message('time scope: ' + self.__time_scope + ' Page: ' + str(page_num) +
                              " Processed, sleeping " + str(sleep_time) + " seconds...", "g")
            if not has_more:
                self.show_message("No More Results!", "g")
                return True
            page_num += 1
            time.sleep(sleep_time)
        return True

    def __parse_keyword_search_result(self, html, buffer):
        """
        parse the response of keyword search
        html: html context of result when respondence contains information
        buffer: tuple, (uname, text_time, text)
        """
        page = etree.HTML(html)
        paragraphs = page.xpath("//p[@class='comment_txt']")
        # contex = open('page.html', 'wb')
        # contex.write(bytes(j, encoding='utf-8'))
        # contex.close()
        times = page.xpath('//a[@class="W_textb"]/@title')
        uids = page.xpath('//a[@class="W_textb"]/@href')
        uid = ''
        uname = ''
        text = ''
        text_time = ''
        idx = 0
        uid_pattern = re.compile('/[0123456789]+/')
        for para in paragraphs:
            uname = str(para.attrib['nick-name'])
            uid = uid_pattern.search(uids[idx]).group(0)[1:-1]
            # self.get_personal_info(uid)
            text = para.xpath('string(.)').strip()
            text_time = times[idx]
            idx += 1
            # filter data
            if (uid + text_time) not in self.filter:
                self.filter[uid + text_time] = True
                # get user info
                buffer.append((uname, text_time, text))

    def get_personal_info(self, uid):
        """(WeiboCrawler, str)

        :param uid:id of the weibo account
        :return:
        """
        headers = {
            "User-Agent":
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36"
        }
        source_url = "https://weibo.cn/%s/info" % uid
        req = urllib.request.Request(url=source_url, headers=headers)
        html = urllib.request.urlopen(req).read()
        page = etree.HTML(html)
        # 获取标签里的所有text()
        text1 = b";".join(page.xpath(b'body/div[@class="c"]/text()'))
        nickname = re.findall(b'\u6635\u79f0[:|\uff1a](.*?);', text1)  # 昵称
        gender = re.findall(b'\u6027\u522b[:|\uff1a](.*?);', text1)  # 性别
        place = re.findall(
            b'\u5730\u533a[:|\uff1a](.*?);', text1)  # 地区（包括省份和城市）
        signature = re.findall(b'\u7b80\u4ecb[:|\uff1a](.*?);', text1)  # 个性签名
        birthday = re.findall(b'\u751f\u65e5[:|\uff1a](.*?);', text1)  # 生日
        sexorientation = re.findall(
            b'\u6027\u53d6\u5411[:|\uff1a](.*?);', text1)  # 性取向
        marriage = re.findall(
            b'\u611f\u60c5\u72b6\u51b5[:|\uff1a](.*?);', text1)  # 婚姻状况
        print(nickname)
        print(gender)
        print(place)
        print(birthday)
        print(sexorientation)
        print(marriage)

    def show_message(self, message, color):
        """

        :param message: str
        :param color: 'g' or 'r'
        :return:
        """
        print(message)

    def save_data(self, data):
        """(WeiboCrawler, [(str, str)])
        data :[(id, weibo)]
        save data as excel file
        """
        total_num = len(data)
        if os.path.exists(self.__file_name):
            wb = openpyxl.load_workbook(self.__file_name)
            value = wb.active['A1'].value
            if len(value) < 6:
                pre_num = 0
            else:
                pre_num = int(wb.active['A1'].value[4:-1])
        else:
            wb = openpyxl.Workbook()
            pre_num = 0

        ws = wb.active
        ws['A1'] = '总共: ' + str(total_num + pre_num) + '条'
        ws['B1'] = '关键字: ' + self.__keyword
        ws['C1'] = '行政区域: ' + self.__region + ":" + self.__subregion
        ws['D1'] = self.__start_time
        ws['E1'] = self.__start_time_curr
        for row in data:
            ws.append(row)
        wb.save(self.__file_name)
        self.show_message("Save data: " + self.__file_name, "g")

    def download(self):
        """
        download data
        """
        data = []
        while self.next_time_scope():
            while not self.craw_data(data):
                time.sleep(3)
            self.save_data(data)
            data = []


def main():
    f = open("./config.yaml", encoding='utf-8')
    config = yaml.load(f)
    f.close()
    login = WeiboLogin.WeiboLogin(config["username"], config["password"])

    # if "file_name" in config:
    #     crawler.set_save_file_name(config["file_name"])

    crawler = Crawler()
    crawler.set_keyword(config["keyword"])
    crawler.set_start_time(config["start_time"])
    crawler.set_stop_time(config["stop_time"])
    crawler.set_weibo_type(config["weibo_type"])
    crawler.set_weibo_containing(config["weibo_containing"])
    crawler.set_region(config["region"])
    crawler.set_subregion(config["subregion"])
    file_name = config["region"] + config["keyword"] + \
        config["start_time"] + "_" + config["stop_time"] + ".xlsx"

    crawler.set_save_file_name(file_name)
    crawler.login = login
    cookie_filename = crawler.get_save_dir() + os.sep + "Cookie.txt"
    file_cookiejar = None
    if os.path.exists(cookie_filename):
        print("Load cookie from: " + cookie_filename)
        file_cookiejar = http.cookiejar.LWPCookieJar()
        file_cookiejar.load(cookie_filename, True, True)
    else:
        login.login()
        file_cookiejar = login.get_cookiejar()
        file_cookiejar.save(cookie_filename, True, True)
        print("Save cookie to: " + cookie_filename)
    crawler.set_cookiejar(file_cookiejar)
    crawler.download()


if __name__ == "__main__":
    main()
